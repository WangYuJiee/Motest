import time
from contextlib import contextmanager
from datetime import datetime
import xlrd
import openpyxl
import json
import requests
from utility.JsbnRSA import JsbnRSA
from config.api_url import GLOBLE_URL


# from JsbnRSA import JsbnRSA
from utility.local_storage import local_data


def get_this_time(UTC_TIME=True):
    """
    获取当前的UTC_TIME
    :param UTC_TIME:
    :return:
    """
    # this_time = datetime.now().strftime("%Y%m%d %H:%M:%S")
    if UTC_TIME:
        # 获取当前的 UTC时间 -- datetime类型值
        this_time = datetime.utcnow()
    else:
        # 获取当前时区的时间 -- datetime类型值
        this_time = datetime.now()
    return datetime.fromtimestamp(time.mktime(this_time.timetuple()))


# def res_check(res, res_check):
#     '''
#     比对接口返回值是否和预期结果一致，一致则返回“pass”，不一致则返回“fail”
#     :param res: 接口返回数据
#     :param res_check: 检查值
#     :return: 测试结果通过或不通过
#     # '''
#     # res = json.dumps(res, ensure_ascii=False)  # 字典转化为字符串
#     # res = res.replace('": "', "=").replace('": ', "=")  # 格式转换为a=1,b=2这样，方便与检查点匹配
#     # res_check = res_check.split(',')
#
#     for r in res_check:
#         if r in res:
#             pass
#         else:
#             return 'fail'
#         return 'pass'
#
# def updateExcel(excel_path,result_list,responses_list):
#     book = openpyxl.load_workbook(excel_path)
#     sheet = book.active
#     for col7 in range(len(result_list)):
#         sheet.cell(col7+2,7,result_list[col7])
#     for col8 in range(len(responses_list)):
#         sheet.cell(col8+2,8,json.dumps(responses_list[col8],ensure_ascii=False))
#     book.save(excel_path)

# 全局token, cookie

cookies = None
Token = ""


def login(username, pwd):
    """
    测试登录
    获取cookie对password加密
    """
    global session, cookie
    session = requests.session()
    try:

        r = session.get(GLOBLE_URL + '/user/session_login')
    except Exception as e:
        raise Exception("Session login error")
    cookies = r.cookies
    cookie = next(
        (cookie for cookie in r.cookies
         if cookie.name == 'public_key'), None
    )
    if not cookie:
        raise ValueError('error cookie')
    public_key = cookie.value
    public_key = json.loads(public_key)
    public_key = public_key.strip('"').replace('\\n', '\n')
    password = JsbnRSA(public_key, is_public=True).encrypt(pwd)

    body = {
        "username": username,
        "password": password
    }
    result = session.post(GLOBLE_URL + '/user/login', json=body, headers={
        'X-From': 'oooo.com',
        'X-Language': 'zh-CN'
    })
    return result


def logout():
    session.post(GLOBLE_URL + '/user/logout')
    return "success"


@contextmanager
def login_decorator():
    try:
        r = login("RS74", "123456")
        global Token
        if r:
            Token = r.json()['token']
            # 设置全局变量
            local_data.__setattr__("token", Token)
        yield
    except Exception as e:
        print("error: ", e)
    finally:
        pass

# if __name__ == '__main__':
#
#     r = login("luxu99","111111")
#     print(r.text)
#     logout()
